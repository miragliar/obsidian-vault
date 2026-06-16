#!/usr/bin/env python3
"""
pbi_usage_report.py — READ-ONLY Power-BI-Tenant-Auswertung:
ALLE semantischen Modelle (Datasets) + Berichte pro Workspace, verknüpft mit der
Report-Nutzung der letzten N Tage aus dem ADMIN-Activity-Log (Distinct User,
Aufrufe, letzter Zugriff). Markiert Lösch-/Prüfkandidaten (0 / wenig Nutzung),
um die Maintenance zu optimieren. Ausgabe: Excel (+ JSON-Summary).

Voraussetzungen
  - Fabric/Power-BI-Admin im Ziel-Tenant (für /admin/activityevents + /admin/reports).
  - openpyxl im venv (pip install openpyxl).
  - Activity-Log reicht systembedingt nur ~30 Tage zurück (Microsoft-Limit).

Auth (Device-Code, 2-phasig, damit der Code ohne Blockieren angezeigt werden kann):
  ./.venv/bin/python pbi_usage_report.py --tenant dobi.ch --auth-init
  ./.venv/bin/python pbi_usage_report.py --tenant dobi.ch --run \
        --out "/abs/pfad/DOBI_Nutzungsanalyse.xlsx" --summary /tmp/dobi_summary.json --days 30
"""
import argparse, json, os, sys, time
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone

import msal
import requests

CID = "04b07795-8ddb-461a-bbee-02f9e1bf7b46"          # Azure-CLI Public-Client (Device-Code)
SCOPE = ["https://analysis.windows.net/powerbi/api/.default"]
BASE = "https://api.powerbi.com/v1.0/myorg"
FLOW_FILE = "/tmp/dobi_pbi_flow.json"


def log(*a):
    print(*a, file=sys.stderr, flush=True)


# ----------------------------------------------------------------------------- auth
def app_for(tenant):
    from auth_common import build_pbi_cache  # Token-Cache verschlüsselt im macOS-Keychain
    return msal.PublicClientApplication(
        CID, authority=f"https://login.microsoftonline.com/{tenant}",
        token_cache=build_pbi_cache(tenant))


def silent(app):
    for acc in app.get_accounts():
        r = app.acquire_token_silent(SCOPE, account=acc)
        if r and "access_token" in r:
            return r["access_token"]
    return None


def auth_init(tenant):
    app = app_for(tenant)
    if silent(app):
        print("AUTH_OK_CACHED")
        return 0
    flow = app.initiate_device_flow(scopes=SCOPE)
    if "user_code" not in flow:
        log("Device-Flow-Fehler:", flow)
        return 2
    json.dump(flow, open(FLOW_FILE, "w"))
    print("DEVICE_URL=" + flow.get("verification_uri", "https://microsoft.com/devicelogin"))
    print("DEVICE_CODE=" + flow.get("user_code", ""))
    print(flow.get("message", ""))
    return 0


CODE_FILE = "/tmp/dobi_device_code.txt"


def get_token(tenant):
    """Silent aus Keychain; sonst Device-Flow IM SELBEN PROZESS (init + poll),
    damit MSAL das volle Zeitfenster pollt. Der Code wird sofort in CODE_FILE
    geschrieben, bevor das (blockierende) Polling startet."""
    app = app_for(tenant)
    t = silent(app)
    if t:
        log("Token aus Keychain-Cache (kein Login nötig).")
        return t
    flow = app.initiate_device_flow(scopes=SCOPE)
    if "user_code" not in flow:
        log("Device-Flow-Fehler:", flow)
        sys.exit(2)
    with open(CODE_FILE, "w") as f:
        f.write("DEVICE_URL=" + flow.get("verification_uri", "https://microsoft.com/devicelogin") + "\n")
        f.write("DEVICE_CODE=" + flow.get("user_code", "") + "\n")
        if flow.get("verification_uri_complete"):
            f.write("DEVICE_URL_COMPLETE=" + flow["verification_uri_complete"] + "\n")
        f.write(flow.get("message", "") + "\n")
    log("DEVICE_CODE=" + flow.get("user_code", ""))
    log(f"(Login-Fenster: expires_in={flow.get('expires_in')}s, interval={flow.get('interval')}s)")
    # Eigenes Poll-Limit — unabhängig von MSALs Default-exit_condition (die hier
    # sofort griff und nach dem ersten 'authorization_pending' abbrach).
    deadline = time.time() + 840
    r = app.acquire_token_by_device_flow(flow, exit_condition=lambda f: time.time() > deadline)
    if not r or "access_token" not in r:
        log("Device-Login fehlgeschlagen:", (r or {}).get("error_description"))
        sys.exit(4)
    return r["access_token"]


# ----------------------------------------------------------------------------- api
class Api:
    def __init__(self, tok):
        self.s = requests.Session()
        self.s.headers["Authorization"] = f"Bearer {tok}"

    def raw(self, path_or_url):
        url = path_or_url if path_or_url.startswith("http") else BASE + path_or_url
        last = None
        for _ in range(6):
            r = self.s.get(url, timeout=120)
            last = r
            if r.status_code == 429:
                time.sleep(int(r.headers.get("Retry-After", "25")))
                continue
            return r
        return last

    def value(self, path):
        """GET .value mit $skip-Paging (admin-Listen)."""
        out, url = [], path
        while url:
            r = self.raw(url)
            if r.status_code in (401, 403):
                raise PermissionError(f"{r.status_code} {url}")
            if r.status_code >= 400:
                log(f"WARN {r.status_code} {url}: {r.text[:160]}")
                break
            j = r.json()
            out += j.get("value", [])
            url = j.get("@odata.nextLink")
        return out

    def activity(self, start, end):
        """/admin/activityevents — Daten liegen unter 'activityEventEntities';
        weiterblättern über 'continuationUri', bis dieser null ist."""
        out, url = [], f"/admin/activityevents?startDateTime='{start}'&endDateTime='{end}'"
        while url:
            r = self.raw(url)
            if r.status_code in (401, 403):
                raise PermissionError(f"{r.status_code} activityevents")
            if r.status_code >= 400:
                log(f"WARN {r.status_code} activity {start[:10]}: {r.text[:120]}")
                break
            j = r.json()
            out += j.get("activityEventEntities", [])
            url = j.get("continuationUri")
        return out


# ----------------------------------------------------------------------------- inventory
def pull_inventory(api):
    groups = api.value("/admin/groups?$top=5000")
    ws_name = {g["id"]: g.get("name", "?") for g in groups}
    ws_type = {g["id"]: g.get("type", "") for g in groups}
    reports = api.value("/admin/reports")
    datasets = api.value("/admin/datasets")
    return ws_name, ws_type, reports, datasets


# ----------------------------------------------------------------------------- activity
def daterange(days):
    now = datetime.now(timezone.utc)
    today = now.date()
    for i in range(days):
        d = today - timedelta(days=i)
        start = f"{d.isoformat()}T00:00:00.000Z"
        if d == today:
            end = now.strftime("%Y-%m-%dT%H:%M:%S.000Z")
        else:
            end = f"{d.isoformat()}T23:59:59.999Z"
        yield d, start, end


def pull_activity(api, days):
    """Activity-Log der letzten `days` Tage aggregieren:
       - agg:     ViewReport pro ReportId (für das Berichte-Sheet)
       - ds_act:  pro DatasetId pro Aktivitätstyp (count, distinct users, last) →
                  erfasst Modell-Nutzung via Bericht, Analyze-in-Excel, DAX-Query u. a.
       - act_hist: Häufigkeit aller Aktivitätstypen (Transparenz/Diagnose)."""
    agg = defaultdict(lambda: {"views": 0, "users": set(), "users7": set(),
                               "last": None, "name": None, "ws": None, "dataset": None})
    ds_act = defaultdict(lambda: defaultdict(lambda: {"count": 0, "users": set(), "last": None}))
    act_hist = Counter()
    now = datetime.now(timezone.utc)
    covered, total_events, view_events = [], 0, 0
    for d, start, end in daterange(days):
        try:
            evs = api.activity(start, end)
        except PermissionError:
            raise
        except Exception as e:  # einzelne Tage robust überspringen
            log(f"WARN activity {d}: {e}")
            continue
        covered.append(d.isoformat())
        total_events += len(evs)
        for e in evs:
            act = e.get("Activity") or "?"
            act_hist[act] += 1
            uid = e.get("UserId") or e.get("UserKey")
            ct = e.get("CreationTime")
            did = e.get("DatasetId")
            if did:
                b = ds_act[did][act]
                b["count"] += 1
                if uid:
                    b["users"].add(uid)
                if ct and (not b["last"] or ct > b["last"]):
                    b["last"] = ct
            if act == "ViewReport":
                rid = e.get("ReportId") or e.get("ArtifactId")
                if not rid:
                    continue
                view_events += 1
                a = agg[rid]
                a["views"] += 1
                if uid:
                    a["users"].add(uid)
                if ct:
                    if not a["last"] or ct > a["last"]:
                        a["last"] = ct
                    try:
                        cdt = datetime.fromisoformat(ct.replace("Z", "+00:00"))
                        if (now - cdt).days < 7 and uid:
                            a["users7"].add(uid)
                    except Exception:
                        pass
                a["name"] = a["name"] or e.get("ReportName")
                a["ws"] = a["ws"] or e.get("WorkspaceName") or e.get("WorkSpaceName")
                a["dataset"] = a["dataset"] or e.get("DatasetId")
    return agg, ds_act, {"covered_days": sorted(covered), "total_events": total_events,
                         "view_events": view_events, "act_hist": dict(act_hist)}


# ----------------------------------------------------------------------------- excel
def short_dt(s):
    return (s or "")[:16].replace("T", " ")


def build_excel(out, ws_name, reports, datasets, agg, days, meta, ws_type=None, ds_act=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    ws_type = ws_type or {}
    ds_act = ds_act or {}
    # Excel-Pivot / externe Tools: bei DOBI über 'AnalyzedByExternalApplication'
    # (Dataset von externer App/Excel abgefragt) + 'AnalyzeInExcel'.
    # 'ConnectFromExternalApplication' trägt KEINE DatasetId → nicht pro Modell zuordenbar.
    EXCEL_ACTS = {"AnalyzeInExcel", "AnalyzedByExternalApplication"}
    QUERY_ACTS = {"ExecuteQueries"}        # REST-DAX (bei DOBI 0) — fürs Total mitgezählt
    SYS_NAMES = {"Usage Metrics Report", "Report Usage Metrics Report",
                 "Dashboard Usage Metrics Report"}

    def ds_signal(did, acts):
        """Summe count + Distinct users + max last über eine Aktivitäts-Menge."""
        cnt, users, last = 0, set(), None
        for act, b in (ds_act.get(did, {}) or {}).items():
            if act in acts:
                cnt += b.get("count", 0)
                users.update(b.get("users", []))
                bl = b.get("last")
                if bl and (not last or bl > last):
                    last = bl
        return cnt, len(users), last

    def category(name, wsid):
        n = (name or "").strip()
        if n in SYS_NAMES or n.startswith("[App] "):
            return "System/Auto"
        wn = (ws_name.get(wsid, "") or "").lower()
        if ws_type.get(wsid, "") in ("PersonalGroup", "Personal") or wn.startswith("personalworkspace"):
            return "Persönlich"
        return "Inhalt"

    RED = PatternFill("solid", fgColor="F4CCCC")
    ORANGE = PatternFill("solid", fgColor="FCE5CD")
    YELLOW = PatternFill("solid", fgColor="FFF2CC")
    GREEN = PatternFill("solid", fgColor="D9EAD3")
    GREY = PatternFill("solid", fgColor="EFEFEF")
    HEAD = PatternFill("solid", fgColor="38761D")
    HFONT = Font(bold=True, color="FFFFFF")

    def reco_report(views, users):
        if views == 0:
            return f"🔴 Löschkandidat – 0 Aufrufe in {days} T", RED
        if users <= 1 and views <= 3:
            return "🟠 Prüfen – kaum genutzt", ORANGE
        if users <= 2 and views < 10:
            return "🟡 Beobachten", YELLOW
        return "🟢 Aktiv", GREEN

    # ---- report rows -------------------------------------------------------
    rrows = []
    for r in reports:
        rid = r.get("id")
        wsid = r.get("workspaceId")
        u = agg.get(rid)
        views = u["views"] if u else 0
        users = len(u["users"]) if u else 0
        users7 = len(u["users7"]) if u else 0
        last = u["last"] if u else None
        rrows.append({
            "ws": ws_name.get(wsid, r.get("workspaceId") or "—"),
            "name": r.get("name", "?"),
            "type": r.get("reportType", "PowerBIReport"),
            "cat": category(r.get("name"), wsid),
            "dataset": r.get("datasetId"),
            "views": views, "users": users, "users7": users7, "last": last,
            "modified": r.get("modifiedDateTime"), "modby": r.get("modifiedBy"),
        })
    ds_name = {d.get("id"): d.get("name", "?") for d in datasets}
    for row in rrows:
        row["dataset_name"] = ds_name.get(row["dataset"], "—" if not row["dataset"] else row["dataset"])
    rrows.sort(key=lambda x: (x["cat"] != "Inhalt", x["views"], x["users"], x["ws"], x["name"].lower()))

    # ---- dataset rows ------------------------------------------------------
    by_ds = defaultdict(lambda: {"views": 0, "users": set(), "reports": 0, "last": None})
    for r in reports:
        did = r.get("datasetId")
        if not did:
            continue
        u = agg.get(r.get("id"))
        b = by_ds[did]
        b["reports"] += 1
        if u:
            b["views"] += u["views"]
            b["users"].update(u["users"])
            if u["last"] and (not b["last"] or u["last"] > b["last"]):
                b["last"] = u["last"]

    drows = []
    for d in datasets:
        did = d.get("id")
        b = by_ds.get(did, {"views": 0, "users": set(), "reports": 0, "last": None})
        xl_cnt, xl_usr, xl_last = ds_signal(did, EXCEL_ACTS)
        q_cnt, q_usr, q_last = ds_signal(did, QUERY_ACTS)
        last_all = max([x for x in (b["last"], xl_last, q_last) if x], default=None)
        drows.append({
            "ws": ws_name.get(d.get("workspaceId"), "—"),
            "name": d.get("name", "?"),
            "owner": d.get("configuredBy") or "—",
            "reports": b["reports"], "views": b["views"], "users": len(b["users"]),
            "xl_cnt": xl_cnt, "xl_usr": xl_usr, "q_cnt": q_cnt, "q_usr": q_usr,
            "last": last_all, "total_signal": b["views"] + xl_cnt + q_cnt,
        })
    drows.sort(key=lambda x: (x["total_signal"] > 0, x["total_signal"], x["name"].lower()))

    wb = Workbook()

    # ---- sheet 1: Übersicht -----------------------------------------------
    ws = wb.active
    ws.title = "Übersicht"
    n_rep = len(rrows)
    n_rep_content = sum(1 for r in rrows if r["cat"] == "Inhalt")
    n_zero = sum(1 for r in rrows if r["views"] == 0)
    n_zero_content = sum(1 for r in rrows if r["cat"] == "Inhalt" and r["views"] == 0)
    n_low = sum(1 for r in rrows if 0 < r["views"] and (r["users"] <= 1 and r["views"] <= 3))
    n_ds = len(drows)
    n_ds_orphan = sum(1 for d in drows if d["reports"] == 0)
    n_ds_unused = sum(1 for d in drows if d["reports"] > 0 and d["views"] == 0)
    n_ds_excel = sum(1 for d in drows if d["xl_cnt"] > 0)
    n_ds_query = sum(1 for d in drows if d["q_cnt"] > 0)
    n_ds_no_signal = sum(1 for d in drows if d["total_signal"] == 0)
    n_ds_no_signal_wrep = sum(1 for d in drows if d["total_signal"] == 0 and d["reports"] > 0)
    ah = meta["activity"].get("act_hist", {})
    cov = meta["activity"]["covered_days"]
    span = f"{cov[0]} … {cov[-1]}" if cov else "—"
    ws.append(["DOBI — Power-BI Nutzungsanalyse: Berichte & semantische Modelle"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    info = [
        ("Erstellt am", meta["generated"]),
        ("Tenant", meta["tenant"]),
        ("Auswertungsfenster (Tage)", days),
        ("Tatsächlich abgedeckt", span),
        ("Workspaces", len(ws_name)),
        ("Semantische Modelle (Datasets)", n_ds),
        ("Berichte gesamt", n_rep),
        ("  davon echte Inhalts-Berichte", n_rep_content),
        ("  davon System/Auto + persönlich", n_rep - n_rep_content),
        ("🔴 Inhalts-Berichte mit 0 Aufrufen (Löschkandidaten)", n_zero_content),
        ("   alle Berichte mit 0 Aufrufen (inkl. System/pers.)", n_zero),
        ("🟠 Berichte kaum genutzt (≤1 User, ≤3 Aufrufe)", n_low),
        ("🟠 Modelle ohne Bericht (Upstream/Composite? prüfen)", n_ds_orphan),
        ("Modelle mit Excel-/Externe-Tool-Nutzung", n_ds_excel),
        ("🔴 Modelle ohne JEDE Nutzung (Bericht+Excel+Query=0)", n_ds_no_signal),
        ("   davon mit ≥1 Bericht (echte Modell-Kandidaten)", n_ds_no_signal_wrep),
        ("ViewReport-Events im Fenster", meta["activity"]["view_events"]),
        ("Excel/Ext.-Abfragen (AnalyzedByExternalApp + AnalyzeInExcel)",
         ah.get("AnalyzedByExternalApplication", 0) + ah.get("AnalyzeInExcel", 0)),
    ]
    for k, v in info:
        ws.append([k, v])
    ws.append([])
    ws.append(["Hinweise / Methodik"])
    ws.cell(ws.max_row, 1).font = Font(bold=True)
    for line in [
        "• Nutzung = ViewReport-Ereignisse aus dem Power-BI Admin-Activity-Log (Audit).",
        "• Distinct User = unterschiedliche Benutzer, die den Bericht im Fenster geöffnet haben.",
        "• Kategorie: 'Inhalt' = echte (geteilte) Berichte · 'Persönlich' = persönliche",
        "  Workspaces Einzelner · 'System/Auto' = von Power BI generiert. Löschfokus = 'Inhalt'.",
        "• Das Admin-Log reicht systembedingt nur ~30 Tage zurück (Microsoft-Limit).",
        "  Für 90 Tage / Trend müsste das Log laufend archiviert werden (täglicher Export).",
        "• 'Letzter Zugriff' bezieht sich nur auf das abgedeckte Fenster.",
        "• Modell-Nutzung erfasst mehrere Kanäle: Bericht-Aufrufe (ViewReport) sowie",
        "  Excel-Pivot / externe Tools (AnalyzedByExternalApplication + AnalyzeInExcel;",
        "  z. B. Excel-Live-Verbindung, DAX Studio). REST-DAX (ExecuteQueries) war im",
        "  Fenster 0. 'ConnectFromExternalApplication' trägt keine DatasetId und ist",
        "  daher nicht pro Modell zuordenbar.",
        "• Rest-Unschärfe: Modelle, die NUR als Upstream eines Composite-Modells dienen,",
        "  lösen evtl. kein eigenes Event aus und können trotzdem 'ohne Nutzung' erscheinen —",
        "  bei DOBI (viele Composite-Cubes) relevant. Modelle weiter fachlich prüfen, nicht",
        "  allein anhand dieser Spalten löschen.",
        "• Reine Lese-Auswertung — es wurde nichts im Tenant verändert.",
        "• Empfehlung ist ein Vorschlag, KEINE automatische Löschung — bitte fachlich prüfen",
        "  (z. B. Quartals-/Jahresberichte werden selten, aber wichtig genutzt).",
    ]:
        ws.append([line])
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 40

    # ---- sheet 2: Berichte – Nutzung --------------------------------------
    s2 = wb.create_sheet("Berichte – Nutzung")
    cols2 = ["Workspace", "Bericht", "Typ", "Kategorie", "Semantisches Modell",
             f"Aufrufe ({days}T)", f"Distinct User ({days}T)", "User (7T)",
             "Letzter Zugriff", "Zuletzt geändert", "Empfehlung"]
    s2.append(cols2)
    for r in rrows:
        if r["cat"] != "Inhalt":
            reco = ("ℹ️ System/Auto – von Power BI verwaltet" if r["cat"] == "System/Auto"
                    else "ℹ️ persönlicher Workspace")
            fill = GREY
        else:
            reco, fill = reco_report(r["views"], r["users"])
        s2.append([r["ws"], r["name"], r["type"], r["cat"], r["dataset_name"],
                   r["views"], r["users"], r["users7"],
                   short_dt(r["last"]) or "—", short_dt(r["modified"]) or "—", reco])
        for c in range(1, len(cols2) + 1):
            s2.cell(s2.max_row, c).fill = fill

    # ---- sheet 3: Semantische Modelle -------------------------------------
    s3 = wb.create_sheet("Semantische Modelle")
    cols3 = ["Workspace", "Semantisches Modell", "Owner", "# Berichte",
             f"Bericht-Aufrufe ({days}T)", "User (Bericht)",
             "Excel/Ext.-Tools Aufrufe", "Excel/Ext.-Tools User",
             "Letzter Zugriff (gesamt)", "Empfehlung"]
    s3.append(cols3)
    for d in drows:
        used = [t for t, on in (("Bericht", d["views"] > 0),
                                ("Excel/Ext", d["xl_cnt"] > 0),
                                ("Query", d["q_cnt"] > 0)) if on]
        if used:
            reco, fill = "🟢 Genutzt (" + "/".join(used) + ")", GREEN
        elif d["reports"] == 0:
            reco, fill = "🟠 Kein Bericht/keine Excel-/Direktnutzung – Upstream/Composite? (prüfen)", ORANGE
        else:
            reco, fill = "🔴 Keine erkennbare Nutzung – prüfen (ggf. Composite-Upstream)", RED
        s3.append([d["ws"], d["name"], d["owner"], d["reports"], d["views"], d["users"],
                   d["xl_cnt"], d["xl_usr"], short_dt(d["last"]) or "—", reco])
        for c in range(1, len(cols3) + 1):
            s3.cell(s3.max_row, c).fill = fill

    # ---- formatting: header, filter, freeze, widths -----------------------
    s2w = [22, 44, 15, 13, 30, 13, 17, 9, 17, 17, 40]
    s3w = [22, 38, 24, 10, 16, 13, 19, 17, 19, 46]
    for sh, ncols, widths in ((s2, len(cols2), s2w), (s3, len(cols3), s3w)):
        for c in range(1, ncols + 1):
            cell = sh.cell(1, c)
            cell.fill = HEAD
            cell.font = HFONT
            cell.alignment = Alignment(vertical="center")
        sh.freeze_panes = "A2"
        sh.auto_filter.ref = f"A1:{get_column_letter(ncols)}{sh.max_row}"
        for i in range(ncols):
            sh.column_dimensions[get_column_letter(i + 1)].width = widths[i] if i < len(widths) else 16

    # ---- sheet 4: Aktivitätstypen (Transparenz / Verifikation) ------------
    s4 = wb.create_sheet("Aktivitätstypen")
    s4.append(["Aktivitätstyp", "Anzahl Events (Fenster)"])
    for k, v in sorted(meta["activity"].get("act_hist", {}).items(), key=lambda x: -x[1]):
        s4.append([k, v])
    for c in (1, 2):
        s4.cell(1, c).fill = HEAD
        s4.cell(1, c).font = HFONT
    s4.freeze_panes = "A2"
    s4.column_dimensions["A"].width = 42
    s4.column_dimensions["B"].width = 24

    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)

    content_zero = sorted([r for r in rrows if r["cat"] == "Inhalt" and r["views"] == 0],
                          key=lambda r: (r["modified"] or ""))
    candidates = [{"ws": r["ws"], "name": r["name"], "modified": short_dt(r["modified"]),
                   "model": r["dataset_name"]} for r in content_zero][:30]
    return {
        "out": out, "n_workspaces": len(ws_name), "n_datasets": n_ds, "n_reports": n_rep,
        "n_reports_content": n_rep_content,
        "n_reports_zero": n_zero, "n_reports_zero_content": n_zero_content, "n_reports_low": n_low,
        "n_datasets_orphan": n_ds_orphan, "n_datasets_unused": n_ds_unused,
        "n_datasets_excel": n_ds_excel, "n_datasets_query": n_ds_query,
        "n_datasets_no_signal": n_ds_no_signal, "n_datasets_no_signal_wrep": n_ds_no_signal_wrep,
        "excel_events": ah.get("AnalyzedByExternalApplication", 0) + ah.get("AnalyzeInExcel", 0),
        "connect_ext_events": ah.get("ConnectFromExternalApplication", 0),
        "query_events": ah.get("ExecuteQueries", 0),
        "window_days": days, "covered": span, "view_events": meta["activity"]["view_events"],
        "top_zero_candidates": candidates,
    }


# ----------------------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tenant", required=True)
    ap.add_argument("--auth-init", action="store_true")
    ap.add_argument("--run", action="store_true")
    ap.add_argument("--out", default="/tmp/DOBI_Nutzungsanalyse.xlsx")
    ap.add_argument("--summary", default="/tmp/dobi_summary.json")
    ap.add_argument("--days", type=int, default=30)
    ap.add_argument("--rawjson", default="/tmp/dobi_raw.json")
    ap.add_argument("--rebuild", help="Excel/Summary aus rawjson neu bauen (keine API-Calls)")
    a = ap.parse_args()

    if a.auth_init:
        sys.exit(auth_init(a.tenant))

    if a.rebuild:
        raw = json.load(open(a.rebuild))
        res = build_excel(a.out, raw["ws_name"], raw["reports"], raw["datasets"],
                          raw["agg"], raw["days"], raw["meta"], raw.get("ws_type", {}),
                          raw.get("ds_act", {}))
        out = {"ok": True, "tenant": raw["meta"].get("tenant"), **res}
        json.dump(out, open(a.summary, "w"), indent=2, ensure_ascii=False)
        print("SUMMARY=" + json.dumps(out, ensure_ascii=False))
        return

    if not a.run:
        ap.error("Entweder --auth-init, --run oder --rebuild angeben.")

    summary = {"ok": False, "tenant": a.tenant}
    try:
        api = Api(get_token(a.tenant))
        log("Inventar (admin) wird geladen …")
        ws_name, ws_type, reports, datasets = pull_inventory(api)
        log(f"  {len(ws_name)} Workspaces · {len(datasets)} Datasets · {len(reports)} Reports")
        log(f"Activity-Log der letzten {a.days} Tage wird geladen …")
        agg, ds_act, act_meta = pull_activity(api, a.days)
        log(f"  {act_meta['total_events']} Events total, davon {act_meta['view_events']} ViewReport, "
            f"über {len(act_meta['covered_days'])} Tage")
        ah = act_meta["act_hist"]
        log(f"  Analyze-in-Excel: {ah.get('AnalyzeInExcel', 0)} · ExecuteQueries: {ah.get('ExecuteQueries', 0)}")
        log("  Top-Aktivitäten: " + ", ".join(f"{k}={v}" for k, v in
            sorted(ah.items(), key=lambda x: -x[1])[:10]))
        meta = {"generated": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
                "tenant": a.tenant, "activity": act_meta}
        raw = {"ws_name": ws_name, "ws_type": ws_type, "reports": reports,
               "datasets": datasets, "days": a.days, "meta": meta,
               "agg": {rid: {**v, "users": sorted(v["users"]), "users7": sorted(v["users7"])}
                       for rid, v in agg.items()},
               "ds_act": {did: {act: {"count": b["count"], "users": sorted(b["users"]),
                                      "last": b["last"]}
                                for act, b in acts.items()}
                          for did, acts in ds_act.items()}}
        json.dump(raw, open(a.rawjson, "w"), ensure_ascii=False)
        log(f"  Rohdaten gecacht: {a.rawjson}")
        res = build_excel(a.out, ws_name, reports, datasets, agg, a.days, meta, ws_type, ds_act)
        summary.update(res)
        summary["ok"] = True
        log(f"✓ Excel geschrieben: {a.out}")
    except PermissionError as e:
        summary["error"] = f"Adminrechte fehlen oder Token ohne Admin-Scope: {e}"
        log("❌ " + summary["error"])
    except SystemExit:
        raise
    except Exception as e:
        summary["error"] = f"{type(e).__name__}: {e}"
        log("❌ " + summary["error"])
    json.dump(summary, open(a.summary, "w"), indent=2, ensure_ascii=False)
    print("SUMMARY=" + json.dumps(summary, ensure_ascii=False))
    sys.exit(0 if summary["ok"] else 5)


if __name__ == "__main__":
    main()
