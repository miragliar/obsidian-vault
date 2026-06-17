#!/usr/bin/env python3
"""Erzeugt 'Bad-Renovation Locarno - Master.xlsx' — konsolidierte Master-Übersicht
mit 10 Tabs (Cover, Kosten, Bauablauf, Kontakte, ToDo, Checkliste, Mail-Chronik,
Risiken, Ausstattung, Dokumente).

Quellen:
- Bestehende Excel-Planung (16.06.2026)
- Obsidian-Notiz Badezimmer-Renovation
- Mail-Korrespondenz Bagno-Koordinationsthread (01.–09.06.2026)
- Dropbox-Inhalt /Bagno matrimoniale/
- Kreditoren-Archiv Badasci

Ziel: alles in EINEM File, ohne Datenredundanz zu verfehlen, klar formatiert,
Schweizer Standard (CHF, DD.MM.YYYY).
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

OUTPUT = Path("/Users/raouleliasmiraglia/Library/CloudStorage/Dropbox/Rechnung/"
              "Locarno Via ai Monti 159A/2026 Rinovazione/"
              "Ristrutturazione Miraglia_Castelli/"
              "Bad-Renovation Locarno - Master.xlsx")

# === Styles ===
THIN = Side(border_style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="305496")  # dunkles Blau
HDR_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SUBHDR_FILL = PatternFill("solid", fgColor="DDEBF7")  # hellblau
SUBHDR_FONT = Font(name="Calibri", size=11, bold=True, color="000000")
TITLE_FONT = Font(name="Calibri", size=18, bold=True, color="1F4E78")
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="595959")
HIGH_FILL = PatternFill("solid", fgColor="FFE699")     # gelb — Achtung
RISK_FILL = PatternFill("solid", fgColor="F8CBAD")     # orange — Risiko
DONE_FILL = PatternFill("solid", fgColor="C6E0B4")     # grün — erledigt
LINK_FONT = Font(name="Calibri", size=11, color="0563C1", underline="single")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def style_header_row(ws, row_idx, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER_ALL
    ws.row_dimensions[row_idx].height = 24


def style_data_rows(ws, start_row, end_row, n_cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_ALL
            if not cell.alignment or cell.alignment.wrap_text is None:
                cell.alignment = LEFT_TOP


def title_block(ws, title, subtitle):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = subtitle
    ws["A2"].font = SUBTITLE_FONT
    ws.row_dimensions[1].height = 28


def add_status_validation(ws, col_letter, start_row, end_row):
    dv = DataValidation(type="list",
                        formula1='"offen,in Arbeit,wartet,erledigt,nicht relevant"',
                        allow_blank=True)
    dv.error = "Bitte aus der Liste wählen"
    dv.errorTitle = "Ungültiger Status"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def add_prio_validation(ws, col_letter, start_row, end_row):
    dv = DataValidation(type="list",
                        formula1='"hoch,mittel,niedrig"',
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


wb = Workbook()
wb.remove(wb.active)

# =======================================================================
# 1) COVER — One-Pager
# =======================================================================
ws = wb.create_sheet("1 · Cover")
title_block(ws,
            "Bad-Renovation Via ai Monti 159a, 6600 Locarno",
            "Wohnung Nr. 10, Residenza Le Palme · Master-Übersicht · Stand 17.06.2026")
set_col_widths(ws, [28, 70])

cover_rows = [
    ("Bauherrschaft", "Elvira & Giovanni Miraglia"),
    ("Objekt", "Hauptbadezimmer (bagno padronale / matrimoniale), Wohnung Nr. 10"),
    ("Adresse", "Via ai Monti 159a, 6600 Locarno-Monti (Condominio Residenza Le Palme)"),
    ("Verwaltung Condominio", "Ing. Gianpiero Forzoni · 091 752 21 38 · 079 583 69 02 · g.forzoni@bluewin.ch"),
    ("Baustart (fix)", "Dienstag, 01.09.2026"),
    ("Geplantes Ende", "ca. Anfang Oktober 2026 (~5 Wochen)"),
    ("Schlüsselübergabe an Bazzi", "Ende Juli 2026"),
    ("Gesamtregie (Lieferant)", "Bazzi Piastrelle SA (Platten + Badeinrichtung)"),
    ("Bauleitung / Koordination",
     "ENTSCHEID 16.06.2026: machen wir SELBST — Bazzi-Aufpreis (~CHF 3'000–4'000) NICHT beauftragt"),
    ("", ""),
    ("Total Bad-Kerngewerke (Bazzi + Badasci)", "CHF 49'531.70 inkl. 8.1 % MwSt"),
    ("davon Bazzi Platten (2026/155)", "CHF 20'783.05"),
    ("davon Bazzi Arredo (2026/156)", "CHF 22'732.90"),
    ("davon Badasci Sanitär (3214)", "CHF 6'015.75"),
    ("Elektro DELO", "noch keine Offerte — telefonisch nachfassen!"),
    ("Maler Pasinelli (26 4021, ganze Whg.)", "CHF 8'939.75 (Bad-Anteil noch ausstehend)"),
    ("", ""),
    ("⚠️ Kritische Punkte (Top 5)",
     "1) DELO antwortet nicht auf Mail — telefonisch · "
     "2) Bauleitung selbst — Bauprogramm von Trunzo (Bazzi) zwingend anfordern · "
     "3) G+E in Türkei während Startwoche (Kappadokien) — Vor-Ort-Vertretung klären · "
     "4) Bazzi-Offerten gegenzeichnen bis 20.07. (sonst keine Materialbestellung) · "
     "5) Strom 230 V für Dusch-WC Geberit AquaClean nicht vergessen"),
    ("", ""),
    ("Dropbox (Offerten + Pläne)",
     "…/Rechnung/Locarno Via ai Monti 159A/2026 Rinovazione/Ristrutturazione Miraglia_Castelli/Bagno matrimoniale/"),
    ("OneDrive (Kreditoren-Rechnungen)",
     "…/Miraglia-BI/Internal/Kreditoren/Archiv/<Firma>/"),
    ("Obsidian-Projektnotiz",
     "40_Resources/scripts/10_Projects/Badezimmer-Renovation Via ai Monti 159a Locarno.md"),
]

r = 4
for label, value in cover_rows:
    ws.cell(row=r, column=1, value=label)
    ws.cell(row=r, column=2, value=value)
    a = ws.cell(row=r, column=1)
    b = ws.cell(row=r, column=2)
    a.font = Font(bold=True, color="1F4E78")
    a.alignment = LEFT_TOP
    b.alignment = LEFT_TOP
    if label.startswith("⚠️"):
        a.fill = HIGH_FILL
        b.fill = HIGH_FILL
    if label.startswith("Bauleitung"):
        b.fill = HIGH_FILL
    if label.startswith("Total Bad"):
        a.font = Font(bold=True, color="C00000")
        b.font = Font(bold=True, color="C00000")
    if label == "":
        ws.row_dimensions[r].height = 6
    else:
        ws.row_dimensions[r].height = max(18, 14 * (1 + len(str(value)) // 70))
    r += 1

ws.sheet_view.showGridLines = False

# =======================================================================
# 2) KOSTEN & BUDGET
# =======================================================================
ws = wb.create_sheet("2 · Kosten & Budget")
title_block(ws, "Kosten & Budget", "Offerten · Zahlungskonditionen · Akonto-Plan")
set_col_widths(ws, [4, 22, 24, 22, 14, 16, 30])

# Tabelle Offerten
hdr = ["#", "Gewerk", "Firma", "Offerte", "Datum", "Betrag CHF", "Bemerkung"]
r = 4
ws.append([])  # leave row 3 free
for i, h in enumerate(hdr, 1):
    ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, len(hdr))

offerten = [
    (1, "Plattenarbeiten", "Bazzi Piastrelle SA", "2026/155", "29.04.2026", 20783.05,
     "Gültigkeit 90 Tage → bis ca. 28.07.2026 ⚠️ vor Ablauf gegenzeichnen"),
    (2, "Badeinrichtung (Lieferung)", "Bazzi Piastrelle SA", "2026/156", "29.04.2026", 22732.90,
     "Lieferzeit Arredo 10–30 Arbeitstage — rechtzeitig bestellen"),
    (3, "Sanitär", "Badasci Fabio Sagl", "3214", "01.05.2026", 6015.75,
     "Preis lt. Cappucci (Bazzi) marktüblich; deliberieren bis 20.07."),
    (4, "Elektro", "Elettricità De Lorenzi (DELO)", "—", "—", None,
     "noch keine Offerte → telefonisch nachfassen (Mossetti / Natasha)"),
    (5, "Maler ganze Whg.", "Pasinelli SA", "26 4021", "30.01.2026", 8939.75,
     "Rechnung 26 4065 (01.04.2026 CHF 11'400) Küche/Salotto bereits fakturiert"),
]
start = r + 1
for row in offerten:
    r += 1
    for i, v in enumerate(row, 1):
        ws.cell(row=r, column=i, value=v)
        if i == 6 and isinstance(v, (int, float)):
            ws.cell(row=r, column=i).number_format = '#,##0.00 "CHF"'
end = r
style_data_rows(ws, start, end, len(hdr))

# Markiere DELO (keine Offerte) gelb
for c in range(1, len(hdr) + 1):
    ws.cell(row=start + 3, column=c).fill = HIGH_FILL

# Total
r += 1
ws.cell(row=r, column=2, value="Total Bad-Kerngewerke (Bazzi 155+156 + Badasci 3214)")
ws.cell(row=r, column=2).font = Font(bold=True)
ws.cell(row=r, column=6, value=20783.05 + 22732.90 + 6015.75)
ws.cell(row=r, column=6).number_format = '#,##0.00 "CHF"'
ws.cell(row=r, column=6).font = Font(bold=True, color="C00000")
for c in range(1, len(hdr) + 1):
    ws.cell(row=r, column=c).border = BORDER_ALL
total_row = r

# Zahlungskonditionen
r += 3
ws.cell(row=r, column=1, value="Zahlungskonditionen")
ws.cell(row=r, column=1).font = SUBHDR_FONT
ws.cell(row=r, column=1).fill = SUBHDR_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
r += 1
zahl_hdr = ["", "Gewerk / Firma", "Anzahlung", "Schlusszahlung", "Ziel", "Bemerkung", ""]
for i, h in enumerate(zahl_hdr, 1):
    ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, len(zahl_hdr))
zahl_start = r + 1
zahl = [
    ("", "Bazzi Arredo (2026/156)", "40 %", "60 %", "30 Tage", "", ""),
    ("", "Bazzi Platten (2026/155)", "90 % laufend", "10 % Schlussrechnung", "10 % nach Abnahme", "Rückbehalt für Mängel", ""),
    ("", "Badasci Sanitär (3214)", "—", "Schluss", "30 Tage netto", "", ""),
    ("", "Pasinelli Maler", "—", "Schluss", "30 Tage netto", "26 4065 bereits gezahlt", ""),
    ("", "DELO Elektro", "—", "Schluss", "30 Tage netto", "Offerte abwarten", ""),
]
for row in zahl:
    r += 1
    for i, v in enumerate(row, 1):
        ws.cell(row=r, column=i, value=v)
style_data_rows(ws, zahl_start, r, len(zahl_hdr))

# Akonto-Plan grobe Projektion
r += 3
ws.cell(row=r, column=1, value="Liquiditäts-Projektion (grob)")
ws.cell(row=r, column=1).font = SUBHDR_FONT
ws.cell(row=r, column=1).fill = SUBHDR_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
r += 1
liq_hdr = ["", "Zeitpunkt", "Position", "Betrag CHF", "kumuliert CHF", "Bemerkung", ""]
for i, h in enumerate(liq_hdr, 1):
    ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, len(liq_hdr))
liq_start = r + 1
# Annahmen: Bazzi-Arredo Anzahlung 40 % vor Lieferung (August)
arredo_az = 22732.90 * 0.40
arredo_rest = 22732.90 - arredo_az
platten_90 = 20783.05 * 0.90
platten_10 = 20783.05 - platten_90
liq = [
    ("", "Aug 2026", "Anzahlung Bazzi Arredo 40 %", arredo_az, arredo_az,
     "Auslöser: gegengezeichnete Offerte 2026/156", ""),
    ("", "Sep 2026", "Bazzi Platten – laufende 90 %", platten_90, arredo_az + platten_90,
     "während Bauphase", ""),
    ("", "Okt 2026", "Bazzi Arredo – Restzahlung 60 %", arredo_rest,
     arredo_az + platten_90 + arredo_rest, "30 Tage nach Lieferung", ""),
    ("", "Okt 2026", "Badasci Schlussrechnung", 6015.75,
     arredo_az + platten_90 + arredo_rest + 6015.75, "nach Endmontage", ""),
    ("", "Nov 2026", "Bazzi Platten – Rückbehalt 10 %", platten_10,
     arredo_az + platten_90 + arredo_rest + 6015.75 + platten_10,
     "nach Mängelbehebung", ""),
    ("", "Okt/Nov 2026", "DELO Elektro (Schätzung)", None, None,
     "Betrag offen — Offerte abwarten", ""),
    ("", "Okt 2026", "Pasinelli Bad-Anteil", None, None,
     "in 26 4021 (CHF 8'939.75) ganze Whg. enthalten", ""),
]
for row in liq:
    r += 1
    for i, v in enumerate(row, 1):
        ws.cell(row=r, column=i, value=v)
        if i in (4, 5) and isinstance(v, (int, float)):
            ws.cell(row=r, column=i).number_format = '#,##0.00 "CHF"'
style_data_rows(ws, liq_start, r, len(liq_hdr))

ws.sheet_view.showGridLines = False
ws.freeze_panes = "A5"

# =======================================================================
# 3) BAUABLAUF
# =======================================================================
ws = wb.create_sheet("3 · Bauablauf")
title_block(ws, "Bauablauf — Wer macht wann was",
            "Soll-Plan ab 01.09.2026. Goldene Regeln: Roh-Installationen VOR Platten · Platten VOR End-Montage · Sanitär demontiert Apparate VOR Bazzi-Abbruch.")
set_col_widths(ws, [4, 26, 18, 20, 38, 32, 12, 16])

hdr = ["#", "Phase", "Zeitfenster", "Firma / Wer", "Tätigkeit", "Vorbedingung", "Lärm?", "Status"]
r = 4
for i, h in enumerate(hdr, 1):
    ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, len(hdr))

ablauf = [
    (0, "Vorbereitung", "Juli / August 2026", "Bauherr (G/E) + Bazzi",
     "Offerten gegenzeichnen; Material bestellen (Arredo-Lieferzeit 10–30 AT!); Schlüssel an Bazzi; Nachbarn informieren; Wasserabstellung mit Verwaltung planen",
     "Auftragserteilung", "—", "in Arbeit"),
    ("1a", "Abbruch Sanitär (ZUERST)", "Wo 1 · Tag 1 · 01.09.", "Badasci",
     "Wasser absperren; Sanitärapparate demontieren (Wanne, WC, Lavabo, Armaturen); Leitungen sichern/verschliessen",
     "Schlüssel bei Bazzi", "teilw.", "offen"),
    ("1b", "Abbruch Platten/Wände", "Wo 1 · ab Tag 1–2 · 01.–05.09.", "Bazzi",
     "Baustelle einrichten & Schutz; alte Platten + Beläge rausspitzen; Cotto-Wand-Demolition für Nischen (Ablage 30×60 + Spiegelschrank 70×150, Tiefe ~10 cm); Schutt entsorgen",
     "Apparate raus (Badasci)", "JA", "offen"),
    (2, "Rohinstallation Sanitär", "Wo 2 · 07.–11.09.", "Badasci",
     "Wasser-/Abwasserleitungen an neue Apparate-Positionen verlegen; ggf. Gebäude-Wasser kurz abstellen",
     "Abbruch fertig", "teilw.", "offen"),
    (2, "Rohinstallation Elektro", "Wo 2 · 07.–11.09.", "DELO",
     "Leitungen/Dosen: Spiegelschrank-LED, STROM FÜR DUSCH-WC (Geberit AquaClean, 230 V!), ggf. zusätzliche Steckdosen",
     "Abbruch fertig; mit Sanitär koordiniert", "teilw.", "offen"),
    (2, "Maurer / Untergrund", "Wo 2 · 07.–11.09.", "Bazzi (Maurer inkl.)",
     "Schlitze schliessen, Untergrund für neue Wanne, Haftbrücke, Glättung — in Bazzi-Offerte enthalten, KEIN separater Auftrag",
     "Roh-Installationen gesetzt", "teilw.", "offen"),
    (3, "Abdichtung", "Wo 2–3 · ab 10.09.", "Bazzi",
     "Estrich-Gefälle Duschtasse; Abdichtung Duschtasse + Wände Duschnische/über Wanne",
     "Untergrund bereit", "nein", "offen"),
    (4, "Plattenarbeiten", "Wo 3 · 14.–18.09.", "Bazzi",
     "Boden + Wände verlegen (Nature Mood / Cementum); Fugen, Edelstahlprofil, Rohrbohrungen, Silikon; Waschtisch-Abdeckplatte",
     "Abdichtung trocken", "JA", "offen"),
    (5, "Sanitär-Endmontage", "Wo 4 · 21.–25.09.", "Badasci",
     "Neue Apparate montieren: Wanne Novellini, WC Geberit, Lavabo Cielo, Armaturen Gessi (bauseits via Bazzi geliefert)",
     "Platten fertig + Apparate geliefert", "nein", "offen"),
    (5, "Badeinrichtung", "Wo 4 · 21.–25.09.", "Bazzi",
     "Duschglas Vismaravetro montieren; Spiegelschrank Keller; Accessoires schwarz matt",
     "Platten fertig", "nein", "offen"),
    (5, "Elektro-Endmontage", "Wo 4 · 21.–25.09.", "DELO",
     "Anschluss Spiegelschrank-LED, Dusch-WC, Steckdosen; Funktionsprüfung",
     "Apparate montiert", "nein", "offen"),
    (6, "Malerarbeiten", "Wo 4–5 · ab 24.09.", "Pasinelli",
     "Nicht verflieste Wandteile verputzen/streichen + Bad-Decke streichen",
     "Platten fertig", "nein", "offen"),
    (7, "Abschluss / Abnahme", "Wo 5 · 28.09.–02.10.", "alle + Bauherr",
     "Silikon final; Funktionsprüfung Wasser/Strom/Whirlpool; Mängelbegehung mit Bazzi; Endreinigung",
     "alle Arbeiten fertig", "nein", "offen"),
]
start = r + 1
for row in ablauf:
    r += 1
    for i, v in enumerate(row, 1):
        ws.cell(row=r, column=i, value=v)
style_data_rows(ws, start, r, len(hdr))
add_status_validation(ws, "H", start, r)

# Lärm-JA Zeilen einfärben
for rr in range(start, r + 1):
    laerm = ws.cell(row=rr, column=7).value
    if laerm == "JA":
        ws.cell(row=rr, column=7).fill = HIGH_FILL

ws.freeze_panes = "A5"
ws.sheet_view.showGridLines = False

# =======================================================================
# 4) KONTAKTE
# =======================================================================
ws = wb.create_sheet("4 · Kontakte")
title_block(ws, "Kontakte — alle Firmen & Personen",
            "Sortiert nach Gewerk. Vor-Ort-Personen separat markiert.")
set_col_widths(ws, [22, 24, 26, 18, 28, 30])

hdr = ["Firma", "Rolle", "Person", "Telefon", "E-Mail", "Bemerkung"]
r = 4
for i, h in enumerate(hdr, 1):
    ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, len(hdr))

kontakte = [
    ("Bazzi Piastrelle SA", "Regie + Showroom / Offerten", "Emiliana Cappucci", "091 785 18 93", "e.cappucci@bazzi.ch",
     "Hauptansprechpartnerin für Offerten 2026/155+156"),
    ("Bazzi Piastrelle SA", "Capo progetto / Bauleitung Lieferant", "Marco Trunzo", "079 774 10 08", "m.trunzo@bazzi.ch",
     "Direktnummer 091 785 18 96 — koordiniert Bauablauf, Materialbestellung"),
    ("Bazzi Piastrelle SA", "Poseur vor Ort", "Maurizio", "—", "—",
     "vor Ort beim Sopralluogo 19.05."),
    ("Bazzi Piastrelle SA", "Allgemein", "info@bazzi.ch", "091 792 16 02", "info@bazzi.ch",
     "Adresse: Via dei Pioppi 10, CP 1245, 6616 Losone"),
    ("Badasci Fabio Sagl", "Inhaber / Sanitär", "Fabio Badasci", "079 458 40 53", "spaions@bluewin.ch",
     "Beim Sopralluogo 19.05. persönlich anwesend"),
    ("Badasci Fabio Sagl", "Monteur vor Ort (Schlüsselträger)", "Giordy Epis", "079 790 67 50", "—",
     "Hat schon bei Cucina geholfen — wichtige Vertrauensperson"),
    ("Badasci Fabio Sagl", "Sitz", "—", "—", "—",
     "Via Montecucco 52A, 6596 Gordola · Suissetec"),
    ("Elettricità De Lorenzi (DELO)", "Büro / Disposition", "Samantha Mossetti / Natasha", "091 751 14 31", "info@delo.ch",
     "🚨 Aktuell keine Antwort auf Bagno-Mail — telefonisch nachfassen!"),
    ("Elettricità De Lorenzi (DELO)", "Elektriker vor Ort", "Maurizio", "079 382 48 08", "—",
     "Hat Cucina-Elektrik gemacht, kennt Wohnung"),
    ("Elettricità De Lorenzi (DELO)", "Geschäftsleitung", "Elia De Lorenzi", "—", "—",
     "Via ai Saleggi 14A, 6600 Locarno"),
    ("Pasinelli SA", "Maler — Büro", "Luisa Voigt", "091 751 77 55", "info@pasinelli.ch",
     "Hauptansprechpartnerin — sehr responsiv"),
    ("Pasinelli SA", "Maler — Familie / Rechnungen", "Nadia Pasinelli", "—", "nadia.pasinelli@pasinelli.ch",
     ""),
    ("Pasinelli SA", "Maler vor Ort", "Dario", "—", "—",
     "Hat Cucina/Salotto gemacht — sehr zufrieden"),
    ("Pasinelli SA", "Sitz", "—", "—", "—",
     "Via D. Galli 4, 6600 Locarno"),
    ("Verwaltung Condominio Le Palme", "Verwaltung / Ing.", "Gianpiero Forzoni", "091 752 21 38 / 079 583 69 02", "g.forzoni@bluewin.ch",
     "Wasserabstellung + Nachbarn-Info"),
    ("3A Bodenbeläge GmbH", "Bodenleger (ganze Whg.)", "Mathias", "—", "info@3ab.ch",
     "Offerte 13050 — Boden ausserhalb Bad"),
    ("Bauherrschaft", "Eigentümer (Vater)", "Giovanni Miraglia", "076 674 33 21", "giovanni@miraglia-bi.com",
     "Hauptkommunikation mit Lieferanten"),
    ("Bauherrschaft", "Eigentümerin (Mutter)", "Elvira Miraglia-Castelli", "076 674 32 55", "elvira@miraglia-bi.com",
     "Administration / cc bei allen Mails"),
    ("Bauherrschaft", "Sohn / IT-Koordination", "Raoul Miraglia", "—", "raoul@miraglia-bi.com",
     "Master-Excel + Obsidian-Dokumentation"),
]
start = r + 1
for row in kontakte:
    r += 1
    for i, v in enumerate(row, 1):
        ws.cell(row=r, column=i, value=v)
style_data_rows(ws, start, r, len(hdr))

# DELO-Zeile hervorheben (keine Antwort)
for rr in range(start, r + 1):
    bem = ws.cell(row=rr, column=6).value or ""
    if "🚨" in bem:
        for c in range(1, len(hdr) + 1):
            ws.cell(row=rr, column=c).fill = HIGH_FILL

ws.freeze_panes = "A5"
ws.sheet_view.showGridLines = False

# =======================================================================
# 5) TODO
# =======================================================================
ws = wb.create_sheet("5 · ToDo")
title_block(ws, "ToDo-Liste — Giovanni & Elvira (+ Raoul)",
            "Status-Spalte hat Dropdown. Prio hat Dropdown.")
set_col_widths(ws, [4, 50, 14, 16, 12, 12, 40])

hdr = ["#", "Aufgabe", "Verantw.", "Bis wann", "Prio", "Status", "Notiz"]
r = 4
for i, h in enumerate(hdr, 1):
    ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, len(hdr))

todos = [
    (1, "Bazzi-Offerten 2026/155 + 2026/156 gegenzeichnen & an Trunzo/Cappucci senden",
     "G", "20.07.2026", "hoch", "offen", "Löst die Materialbestellung aus — keine Bestellung ohne Unterschrift"),
    (2, "Badasci Sanitär-Offerte 3214 (CHF 6'015.75) bestätigen / beauftragen",
     "G", "20.07.2026", "hoch", "offen", ""),
    (3, "DELO TELEFONISCH erreichen: Elektro-Bedarf Bad (Dusch-WC-Strom!, Spiegelschrank) + Offerte + Termin Woche 2",
     "G", "diese Woche", "hoch", "offen", "🚨 Reagiert nicht auf Mail — Samantha Mossetti / Natasha unter 091 751 14 31"),
    (4, "Bauprogramm mit Daten von Bazzi (Trunzo) anfordern",
     "G", "30.06.2026", "hoch", "offen", "Basis für unsere Selbst-Koordination"),
    (5, "Pasinelli (Luisa Voigt): Bad-Leistungsumfang + Termin schriftlich bestätigen",
     "G/E", "Juli 2026", "mittel", "offen", "Sie hat am 02.06. um Detail-Klärung gebeten"),
    (6, "Wandstärke-Frage Nischen mit Bazzi klären (Machbarkeit)",
     "G", "vor Baustart", "mittel", "offen", "Offen seit 29.04. — bei Baubeginn final bestätigen"),
    (7, "Material-/Apparate-Schnittstelle Bazzi ↔ Badasci klären (wer liefert was)",
     "G", "Juli 2026", "mittel", "offen",
     "Badasci montiert nur bauseits Geliefertes — Lücken/Doppelbestellungen vermeiden"),
    (8, "Schlüsselübergabe an Bazzi organisieren",
     "G/E", "Ende Juli 2026", "hoch", "offen", ""),
    (9, "Wasserabstellung mit Verwaltung Forzoni + Nachbarn koordinieren",
     "E", "August 2026", "mittel", "offen", "Forzoni 091 752 21 38"),
    (10, "Nachbarn-Infoblatt verteilen (Cari vicini)",
     "E", "August 2026", "mittel", "offen", "Vorlage in Dropbox vorhanden"),
    (11, "Freitags keine Lärmarbeiten allen Firmen schriftlich bestätigen",
     "G", "Juli 2026", "niedrig", "offen", ""),
    (12, "Erreichbarkeit Startwoche regeln (G+E in Türkei → E als Ansprechpartnerin?)",
     "G/E", "August 2026", "hoch", "offen",
     "🚨 Abbruch ist die heikelste Phase — ohne Entscheider vor Ort riskant"),
    (13, "Tägliche Vor-Ort-Kontaktliste erstellen (Epis, Maurizio-DELO, Maurizio-Bazzi)",
     "G", "August 2026", "niedrig", "offen", "Siehe Tab Kontakte"),
    (14, "Entscheid festhalten: Koordination selbst (kein Bazzi-Aufpreis 3–4k)",
     "G", "16.06.2026", "—", "erledigt", "ERLEDIGT 16.06.2026"),
    (15, "Liquiditäts-/Akontoplan (Bazzi 40/60 + 90/10) pflegen",
     "G", "laufend", "niedrig", "offen", "siehe Tab Kosten & Budget"),
    (16, "Master-Excel mit Mailverkehr-Analyse erstellen",
     "R", "17.06.2026", "—", "erledigt", "Diese Datei — ERLEDIGT"),
]
start = r + 1
for row in todos:
    r += 1
    for i, v in enumerate(row, 1):
        ws.cell(row=r, column=i, value=v)
style_data_rows(ws, start, r, len(hdr))
add_status_validation(ws, "F", start, r)
add_prio_validation(ws, "E", start, r)

# Hoch-Prio rot hinterlegen, erledigt grün
for rr in range(start, r + 1):
    prio = ws.cell(row=rr, column=5).value
    status = ws.cell(row=rr, column=6).value
    if status == "erledigt":
        for c in range(1, len(hdr) + 1):
            ws.cell(row=rr, column=c).fill = DONE_FILL
    elif prio == "hoch":
        ws.cell(row=rr, column=5).fill = HIGH_FILL
        bem = ws.cell(row=rr, column=7).value or ""
        if "🚨" in bem:
            for c in range(1, len(hdr) + 1):
                ws.cell(row=rr, column=c).fill = RISK_FILL

ws.freeze_panes = "A5"
ws.sheet_view.showGridLines = False

# =======================================================================
# 6) CHECKLISTE — kritische Schnittstellen
# =======================================================================
ws = wb.create_sheet("6 · Checkliste")
title_block(ws, "Checkliste — worauf achten & was fragen",
            "Wir übernehmen die Koordination selbst — das hier sind die kritischen Schnittstellen.")
set_col_widths(ws, [22, 22, 42, 38, 12])

hdr = ["Bereich", "An wen", "Frage / Kontrollpunkt", "Warum wichtig", "Erledigt?"]
r = 4
for i, h in enumerate(hdr, 1):
    ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, len(hdr))

check = [
    ("Reihenfolge", "alle Firmen", "Roh-Installationen Sanitär + Elektro VOR den Platten fertig?",
     "Häufigste Fehlerquelle — nachträgliche Schlitze zerstören neue Platten.", ""),
    ("Reihenfolge", "Bazzi / Badasci", "Platten fertig & trocken VOR der Sanitär-Endmontage?",
     "Apparate werden auf die fertigen Platten montiert.", ""),
    ("Material", "Bazzi (Cappucci/Trunzo)",
     "Arredo (Wanne, WC, Armaturen…) bestellt? Liefertermin VOR Woche 4?",
     "Lieferzeit 10–30 AT laut Offerte 2026/156 — sonst Stillstand bei der Endmontage.", ""),
    ("Material", "Bazzi ↔ Badasci",
     "Wer liefert welche Apparate — ist ALLES, was Badasci montiert, in der Bazzi-Lieferung?",
     "Badasci montiert nur bauseits Geliefertes (Offerte 3214). Lücken/Doppelbestellungen vermeiden.", ""),
    ("Elektro", "DELO",
     "Stromanschluss Dusch-WC (Geberit AquaClean, 230V) + Spiegelschrank-LED eingeplant?",
     "Wird oft vergessen — AquaClean braucht Strom; nachträglich teuer und sichtbar.", ""),
    ("Elektro", "DELO",
     "DELO reagiert nicht auf Mail → TELEFONISCH Offerte + Termin (Woche 2) klären.",
     "Einzige Firma ohne Rückmeldung; Elektro liegt auf dem kritischen Pfad.", ""),
    ("Technik", "Bazzi (Trunzo)",
     "Nischen-Tiefe ca. 10 cm laut Vor-Ort-Einschätzung machbar — bei Baubeginn final bestätigen.",
     "Eventualposition: bei Durchbruch zur Gegenseite sind Wiederherstellung + Maler NICHT im Preis.", ""),
    ("Maler", "Pasinelli (Luisa Voigt)",
     "Bad-Leistungsumfang schriftlich bestätigt (Wände + Decke)? Termin nach den Platten?",
     "Bisher nur mündlich umrissen; Pasinelli wartet aufs Bauprogramm.", ""),
    ("Programm", "Bazzi (Trunzo)",
     "Verbindliches Bauprogramm mit Daten pro Gewerk anfordern.",
     "Wir koordinieren selbst → ohne Termine keine Schnittstellen-Steuerung.", ""),
    ("Nachbarn", "Verwaltung Forzoni",
     "Wasserabstellung Datum/Dauer abklären + Nachbarn-Infoblatt verteilen.",
     "Gebäude-Wasseranschluss muss zeitweise zu — Vorankündigungspflicht.", ""),
    ("Nachbarn", "alle Firmen",
     "Freitags keine Lärmarbeiten an alle kommuniziert?",
     "Rücksicht auf Nachbarin im Stock darunter (v.a. Abbruch + Plattenschnitt).", ""),
    ("Organisation", "G + E",
     "ACHTUNG: G + E in der Startphase in der Türkei (Kappadokien) — wer entscheidet vor Ort? "
     "Bazzi-Bauleitung für diese Phase / Start verschieben / Vertrauensperson?",
     "Abbruch ist die heikelste Phase (Überraschungen hinter den Platten) — ohne Entscheider vor Ort riskant.", ""),
    ("Organisation", "G / E",
     "Tägliche Vor-Ort-Kontakte definiert? (Giordy Epis / Maurizio Bazzi)",
     "Selbst-Koordination heisst: pro Tag eine erreichbare Ansprechperson.", ""),
    ("Abnahme", "Bazzi + alle",
     "Gemeinsame Mängelbegehung + Funktionsprüfung am Ende terminiert?",
     "Mängel vor Schlusszahlung dokumentieren (10 % Rückbehalt bei Bazzi-Platten).", ""),
    ("Zahlung", "alle",
     "Akonti/Zahlungen im Blick (Bazzi 40/60 bzw. 90/10; übrige 30 T netto).",
     "Liquiditätsplanung; Akontozahlungen korrekt freigeben.", ""),
]
start = r + 1
for row in check:
    r += 1
    for i, v in enumerate(row, 1):
        ws.cell(row=r, column=i, value=v)
style_data_rows(ws, start, r, len(hdr))
add_status_validation(ws, "E", start, r)

# Kritische Zeilen markieren
for rr in range(start, r + 1):
    frage = (ws.cell(row=rr, column=3).value or "")
    if "DELO reagiert nicht" in frage or "ACHTUNG" in frage:
        for c in range(1, len(hdr) + 1):
            ws.cell(row=rr, column=c).fill = RISK_FILL

ws.freeze_panes = "A5"
ws.sheet_view.showGridLines = False

# =======================================================================
# 7) MAIL-CHRONIK
# =======================================================================
ws = wb.create_sheet("7 · Mail-Chronik")
title_block(ws, "Mail-Chronik — Bagno-Koordination",
            "Chronologie der Bagno-relevanten Mails (Quelle: Postfach giovanni@miraglia-bi.com).")
set_col_widths(ws, [14, 8, 24, 30, 38, 60])

hdr = ["Datum", "Richt.", "Absender", "Empfänger", "Betreff", "Kerninhalt"]
r = 4
for i, h in enumerate(hdr, 1):
    ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, len(hdr))

mails = [
    ("03.02.2026", "AUS", "elvira@miraglia-bi.com",
     "info@pasinelli.ch (cc Giovanni)",
     "AW: Preventivo opere da pittore — Locarno-Monti",
     "Bestätigung Offerte Pasinelli 26 4021 (CHF 8'939.75) für ganze Whg. inkl. spätere Bad-Arbeiten"),
    ("14.02.2026", "AUS", "giovanni@miraglia-bi.com",
     "g.forzoni@bluewin.ch (cc Elvira)",
     "Lavori via ai Monti 159a — Appartamento Miraglia",
     "Vorabinformation an Verwaltung Forzoni über Renovations-Lavori + Liste aller Ditten"),
    ("16.02.2026", "AUS", "giovanni@miraglia-bi.com",
     "g.forzoni, elvira, B.Maas (Kralewski) u.a.",
     "Lavori / Arbeiten — Appartamento 10",
     "Nachbarn-Infoblatt digital verteilt (auf Wunsch Forzoni)"),
    ("24.02.2026", "AUS", "giovanni@miraglia-bi.com",
     "info@delo.ch",
     "Coordinamento lavori elettrici — nuova cucina e salotto",
     "Cucina-Smontage erfolgt, Strom abgeklemmt, Termin Maurizio anfragen"),
    ("25.02.2026", "EIN", "info@delo.ch (Natasha)",
     "giovanni@miraglia-bi.com",
     "R: Coordinamento lavori elettrici",
     "Bestätigt: Maurizio kommt 02.03.2026 um 15:30"),
    ("12.03.–04.05.2026", "↔", "info@pasinelli.ch ↔ giovanni",
     "—", "Preventivo opere da pittore — Locarno-Monti (Thread)",
     "Cucina/Salotto-Arbeiten von Dario erledigt; mehrere Ritocchi-Termine; Pasinelli bestätigt am 04.05. Abschluss + Schlüsselrückgabe"),
    ("20.03.2026", "↔", "giovanni ↔ info@delo.ch (Samantha)",
     "—", "Coordinamento lavori elettrici (Thread)",
     "Maurizio in Urlaub; neuer Termin 27.03. Letzte aktive DELO-Mail vor Bagno-Thema."),
    ("01.04.2026", "EIN", "nadia.pasinelli@pasinelli.ch via Elvira",
     "giovanni@miraglia-bi.com (Fwd)",
     "Opere da pittore eseguite a Locarno Monti",
     "Pasinelli-Rechnung 26 4065 = CHF 11'400 (Küche/Salotto) zugestellt"),
    ("10.05.2026", "AUS", "elvira@miraglia-bi.com",
     "e.cappucci@bazzi.ch (cc Giovanni)",
     "Via ai Monti 159a, Fam Miraglia",
     "Bittet Cappucci um Einschätzung der Badasci-Sanitäroffer"),
    ("11.05.2026", "EIN", "e.cappucci@bazzi.ch",
     "elvira@miraglia-bi.com (cc Giovanni)",
     "R: Via ai Monti 159a, Fam Miraglia",
     "Cappucci: 'Preventivo ha un ottimo prezzo vedendo le altre offerte con lavori simili.'"),
    ("19.05.2026", "Termin", "Sopralluogo vor Ort",
     "Bazzi (Maurizio) + Badasci (Fabio Badasci) + Bauherrschaft",
     "Begehung Wohnung",
     "Klärung Leistungsumfang; Cappucci/Trunzo nicht dabei"),
    ("01.06.2026 21:20", "AUS", "giovanni@miraglia-bi.com",
     "e.cappucci@bazzi.ch, spaions@bluewin.ch, info@pasinelli.ch, info@delo.ch (cc Elvira)",
     "Rinnovo bagno — Via ai Monti 159a, Locarno — coordinamento lavori",
     "INITIAL: Bazzi = Regie; Ziel Aug/Sep; Schlüssel an Bazzi Ende Juli; Fr keine Lärmarbeiten; Wasserabstellung melden; DELO bei Bedarf"),
    ("02.06.2026 09:32", "EIN", "info@pasinelli.ch (Luisa Voigt)",
     "an alle 4 + Bauherr",
     "R: Rinnovo bagno — coordinamento lavori",
     "Bittet um detaillierte Klarstellung Leistungsumfang Pasinelli im Bad; wartet auf Bauprogramm von Bazzi"),
    ("02.06.2026 11:29", "EIN", "m.trunzo@bazzi.ch",
     "an alle 4 + Bauherr",
     "R: Rinnovo bagno — coordinamento lavori",
     "Bazzi kann ab 01.09.2026 intervenieren; wartet auf Delibera ufficiale"),
    ("07.06.2026 23:01", "AUS", "giovanni@miraglia-bi.com",
     "an alle 4 + Bauherr",
     "AW: Rinnovo bagno — coordinamento lavori",
     "01.09.2026 bestätigt; Türkei-Urlaub Startwoche; Schlüssel Juli/Aug; Bauprogramm anfordern; KLÄRUNG Leistungsumfang pro Gewerk; Fr: Arbeit ok, aber ohne Lärm"),
    ("09.06.2026 10:27", "EIN", "m.trunzo@bazzi.ch",
     "an alle 4 + Bauherr",
     "R: Rinnovo bagno — coordinamento lavori",
     "Bestätigt 01.09.; Materialbestellung nach gegengezeichneter Offerte; PRECISAZIONE: Bauleitung = separate Leistung CHF 3'000–4'000 für ein Bad; alternativ Selbst-Koordination → 16.06. ABGELEHNT"),
    ("16.06.2026", "Entscheid", "Bauherrschaft G+E intern", "—", "Entscheid Bauleitung",
     "Wir koordinieren SELBST — Bazzi-Aufpreis 3-4k nicht beauftragt"),
    ("17.06.2026 07:12", "EIN", "giovanni@miraglia-bi.com",
     "raoul@miraglia-bi.com",
     "Noch ein kurzes Video - Obsidian im Einsatz",
     "Auftrag Raoul: Video-Demo Obsidian + Master-Excel-Erstellung"),
]
start = r + 1
for row in mails:
    r += 1
    for i, v in enumerate(row, 1):
        ws.cell(row=r, column=i, value=v)
style_data_rows(ws, start, r, len(hdr))

# Initial-Mail + Precisazione-Mail hervorheben
for rr in range(start, r + 1):
    subj = (ws.cell(row=rr, column=5).value or "")
    inh = (ws.cell(row=rr, column=6).value or "")
    if "PRECISAZIONE" in inh or "ABGELEHNT" in inh:
        for c in range(1, len(hdr) + 1):
            ws.cell(row=rr, column=c).fill = HIGH_FILL
    if "INITIAL" in inh or rr == start + 11:  # 01.06.
        ws.cell(row=rr, column=5).font = Font(bold=True)

ws.freeze_panes = "A5"
ws.sheet_view.showGridLines = False

# =======================================================================
# 8) RISIKEN
# =======================================================================
ws = wb.create_sheet("8 · Risiken")
title_block(ws, "Risikoregister",
            "P = Wahrscheinlichkeit (1–5), A = Auswirkung (1–5), Score = P × A")
set_col_widths(ws, [4, 32, 38, 6, 6, 8, 38, 14])

hdr = ["#", "Risiko", "Beschreibung / Ursache", "P", "A", "Score", "Massnahme", "Status"]
r = 4
for i, h in enumerate(hdr, 1):
    ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, len(hdr))

risks = [
    (1, "DELO antwortet nicht", "DELO im Verteiler, aber keine Antwort auf Mails 01.+07.06. Elektro liegt auf dem kritischen Pfad (Roh-Installation Woche 2).",
     4, 5, None, "Telefonisch (091 751 14 31, Mossetti/Natasha) bis Ende Woche 25; bei Bedarf alternative Firma im Standby (Elettro Mastai SA = Le-Palme-Standardfirma).", "offen"),
    (2, "Niemand vor Ort in Startwoche", "G+E sind in Kappadokien (Türkei) während der heikelsten Phase (Abbruch). Entscheidungen über Cotto-Wand, versteckte Leitungen etc. unmöglich.",
     5, 5, None, "Vertrauensperson (Epis? Trunzo? alternativ Bazzi-Bauleitung punktuell für diese Woche zukaufen) ODER Start verschieben auf 08.09.2026.", "offen"),
    (3, "Materialbestellung verspätet", "Arredo-Lieferzeit 10–30 AT. Bei Auftragsfreigabe nach 20.07. evtl. Lieferung erst Mitte September → Stillstand bei Endmontage.",
     3, 4, None, "Offerten gegenzeichnen bis 20.07. (Bazzi-Offerte gültig 90 T bis 28.07.).", "in Arbeit"),
    (4, "Versteckte Leitungen / Bausubstanz", "Beim Abbruch der Cotto-Wand kann sich herausstellen, dass Nische 10 cm nicht machbar / Wiederherstellung der Gegenseite nötig.",
     3, 3, None, "Vor Baustart Tiefenmessung mit Trunzo; Eventualposition im Budget einplanen; Maler/Maurer-Reserve.", "offen"),
    (5, "Wasserabstellung Konflikt", "Gebäude-Wasser für mehrere Stunden zu — bei kurzfristiger Ankündigung Konflikt mit Nachbarn / Verwaltung.",
     2, 3, None, "Mit Badasci konkretes Datum/Dauer fixieren → Forzoni 2 Wochen vorab informieren → Nachbarn-Infoblatt.", "offen"),
    (6, "Lärm-Beschwerden Nachbarin", "Nachbarin Stockwerk darunter hat um lärmarme Freitage gebeten. Bei Abbruch und Plattenschnitt sehr laut.",
     3, 2, None, "Allen Firmen schriftlich: Fr keine Lärmarbeiten. Lärmtage auf Mo–Do bündeln.", "offen"),
    (7, "Schnittstelle Bazzi ↔ Badasci unklar", "Wer liefert welche Apparate? Badasci montiert nur bauseits Geliefertes; Lücken/Doppelbestellungen kosten Zeit.",
     3, 3, None, "Schriftliche Bestätigung Bazzi: Lieferumfang vollständig? Abgleich mit Badasci-Offerte 3214.", "offen"),
    (8, "Dusch-WC ohne Strom", "Geberit AquaClean braucht 230 V — bei vergessenem Strom­anschluss nachträglich sehr teuer (Wand öffnen).",
     2, 5, None, "Mit DELO explizit klären; im Bauprogramm Woche 2 als Pflichtposition; bei Endmontage prüfen.", "offen"),
    (9, "Mängel ohne Rückbehalt", "Wenn Schlussrechnung vor Mängelfeststellung gezahlt wird, schwer durchsetzbar.",
     2, 3, None, "10 % Rückbehalt bei Bazzi-Platten konsequent nutzen; Mängelbegehung dokumentiert.", "offen"),
    (10, "Pasinelli ohne Termin", "Pasinelli wartet auf Bauprogramm; bei verspätetem Bad-Termin (nach 24.09.) Kollision mit anderen Aufträgen.",
     2, 2, None, "Bauprogramm sofort an Voigt schicken; Termin frühzeitig blocken.", "offen"),
]
start = r + 1
for row in risks:
    r += 1
    n, name, desc, p, a, _, ms, st = row
    ws.cell(row=r, column=1, value=n)
    ws.cell(row=r, column=2, value=name)
    ws.cell(row=r, column=3, value=desc)
    ws.cell(row=r, column=4, value=p)
    ws.cell(row=r, column=5, value=a)
    ws.cell(row=r, column=6, value=f"=D{r}*E{r}")
    ws.cell(row=r, column=7, value=ms)
    ws.cell(row=r, column=8, value=st)
style_data_rows(ws, start, r, len(hdr))
add_status_validation(ws, "H", start, r)

# Score-Färbung
from openpyxl.formatting.rule import ColorScaleRule
ws.conditional_formatting.add(
    f"F{start}:F{r}",
    ColorScaleRule(start_type="num", start_value=1, start_color="C6E0B4",
                   mid_type="num", mid_value=12, mid_color="FFE699",
                   end_type="num", end_value=25, end_color="F8CBAD")
)

ws.freeze_panes = "A5"
ws.sheet_view.showGridLines = False

# =======================================================================
# 9) AUSSTATTUNG & MATERIAL
# =======================================================================
ws = wb.create_sheet("9 · Ausstattung")
title_block(ws, "Geplante Ausstattung & Material",
            "Highlights aus Offerte 2026/156 (Bazzi Arredo) und 2026/155 (Bazzi Platten)")
set_col_widths(ws, [22, 28, 26, 36, 16])

hdr = ["Bereich", "Position", "Marke / Modell", "Spezifikation", "Quelle / Offerte"]
r = 4
for i, h in enumerate(hdr, 1):
    ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, len(hdr))

ausstattung = [
    ("Badewanne", "Whirlpool-Wanne", "Novellini Divina Hydro Plus",
     "190 × 90 cm, Whirlpool + Airpool, Chromolight, Bluetooth", "2026/156"),
    ("Dusche", "Thermostat + Kopfbrause", "Gessi Emporio Shower",
     "Kopfbrause 30 cm", "2026/156"),
    ("Dusche", "Glaswand", "Vismaravetro SK-IN",
     "120 × 200 cm, Profile schwarz matt", "2026/156"),
    ("WC", "Dusch-WC (wandhängend)", "Geberit AquaClean Sela",
     "230 V Stromanschluss erforderlich", "2026/156"),
    ("WC", "Betätigungsplatte", "Geberit Sigma20",
     "schwarz matt", "2026/156"),
    ("Waschbereich", "Aufsatzbecken", "Cielo Le Bacinelle",
     "—", "2026/156"),
    ("Waschbereich", "Armaturen", "Gessi Via Manzoni",
     "schwarz matt", "2026/156"),
    ("Waschbereich", "Unterbau", "Petra",
     "Finish Geowood Mova", "2026/156"),
    ("Spiegel", "Spiegelschrank", "Keller",
     "150 × 73 cm, LED beleuchtet — Strom DELO", "2026/156"),
    ("Boden", "Platten Boden", "Nature Mood",
     "schwarz matt — Look durchgehend", "2026/155"),
    ("Wände", "Platten Wand", "Cementum",
     "schwarz matt — Look durchgehend", "2026/155"),
    ("Wände", "Nische 1", "—",
     "Ablage 30 × 60 cm, Tiefe ~10 cm (Eventualposition)", "2026/155 + Vor-Ort"),
    ("Wände", "Nische 2", "—",
     "Spiegelschrank 70 × 150 cm, Tiefe ~10 cm (Eventualposition)", "2026/155 + Vor-Ort"),
]
start = r + 1
for row in ausstattung:
    r += 1
    for i, v in enumerate(row, 1):
        ws.cell(row=r, column=i, value=v)
style_data_rows(ws, start, r, len(hdr))

ws.freeze_panes = "A5"
ws.sheet_view.showGridLines = False

# =======================================================================
# 10) DOKUMENTE
# =======================================================================
ws = wb.create_sheet("10 · Dokumente")
title_block(ws, "Dokumenten-Verzeichnis",
            "Alle relevanten Dateien (PDF, Pläne, Renderings, Offerten, Rechnungen). Hyperlinks öffnen direkt aus Dropbox/OneDrive.")
set_col_widths(ws, [28, 22, 16, 14, 70])

hdr = ["Dokument", "Typ", "Datum", "Firma", "Pfad"]
r = 4
for i, h in enumerate(hdr, 1):
    ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, len(hdr))

DROP_BAGNO = ("/Users/raouleliasmiraglia/Library/CloudStorage/Dropbox/Rechnung/"
              "Locarno Via ai Monti 159A/2026 Rinovazione/"
              "Ristrutturazione Miraglia_Castelli/Bagno matrimoniale")
DROP_BAGNO_FINAL = DROP_BAGNO + "/Final"
DROP_ROOT = ("/Users/raouleliasmiraglia/Library/CloudStorage/Dropbox/Rechnung/"
             "Locarno Via ai Monti 159A/2026 Rinovazione/"
             "Ristrutturazione Miraglia_Castelli")
OD_BADASCI = ("/Users/raouleliasmiraglia/Library/CloudStorage/"
              "OneDrive-MiragliaBusiness-Intelligence/Miraglia-BI/Internal/"
              "Kreditoren/Archiv/Badasci Fabio SAGL")

docs = [
    ("Offerta 2026-155 Piastrelle.pdf", "Offerte", "29.04.2026", "Bazzi",
     f"{DROP_BAGNO_FINAL}/Offerta 2026-155 Piastrelle.pdf"),
    ("Offerta 2026-156 Arredo Bagno.pdf", "Offerte", "29.04.2026", "Bazzi",
     f"{DROP_BAGNO_FINAL}/Offerta 2026-156 Arredo Bagno.pdf"),
    ("2026-05-01_Badasci_Offerta — via ai Monti.pdf", "Offerte", "01.05.2026", "Badasci",
     f"{OD_BADASCI}/2026-05-01_Badasci Fabio SAGL_Offerta - via ai Monti 159a, Locarno_A-4167.pdf"),
    ("2026-03-20_Badasci_Curva GEBERIT DN 56 45°.pdf", "Rechnung", "20.03.2026", "Badasci",
     f"{OD_BADASCI}/2026-03-20_Badasci Fabio SAGL_Curva GEBERIT DN 56 45°_A-4038.pdf"),
    ("2026-05-10_Badasci_Allacciamento Cucina.pdf", "Rechnung", "10.05.2026", "Badasci",
     f"{OD_BADASCI}/2026-05-10_Badasci Fabio SAGL_Allacciamento Cucina nella vostra casa a Locarno._A-4170.pdf"),
    ("Miraglia - pianta.pdf", "Grundriss", "22.04.2026", "Bazzi/Studio",
     f"{DROP_BAGNO_FINAL}/Miraglia - pianta.pdf"),
    ("Miraglia 1.pdf", "Plan/Rendering", "29.04.2026", "Bazzi/Studio",
     f"{DROP_BAGNO_FINAL}/Miraglia 1.pdf"),
    ("Miraglia 2.pdf", "Plan/Rendering", "29.04.2026", "Bazzi/Studio",
     f"{DROP_BAGNO_FINAL}/Miraglia 2.pdf"),
    ("Miraglia 3.pdf", "Plan/Rendering", "29.04.2026", "Bazzi/Studio",
     f"{DROP_BAGNO_FINAL}/Miraglia 3.pdf"),
    ("Miraglia 4.pdf", "Plan/Rendering", "29.04.2026", "Bazzi/Studio",
     f"{DROP_BAGNO_FINAL}/Miraglia 4.pdf"),
    ("Miraglia - Progetto bagno pianta.pdf", "Grundriss (1. Variante)", "21.03.2026", "Bazzi/Studio",
     f"{DROP_BAGNO}/Miraglia - Progetto bagno pianta.pdf"),
    ("Miraglia - Progetto bagno rendering 1.pdf", "Rendering (1. Variante)", "21.03.2026", "Bazzi/Studio",
     f"{DROP_BAGNO}/Miraglia - Progetto bagno rendering 1.pdf"),
    ("Miraglia - Progetto bagno rendering 2.pdf", "Rendering (1. Variante)", "21.03.2026", "Bazzi/Studio",
     f"{DROP_BAGNO}/Miraglia - Progetto bagno rendering 2.pdf"),
    ("Bad-Renovation Locarno - Planung und ToDo.xlsx", "Planung (Vorgänger)", "16.06.2026", "Bauherr",
     f"{DROP_ROOT}/Bad-Renovation Locarno - Planung und ToDo.xlsx"),
    ("Bad-Renovation Locarno - Master.xlsx", "Master (diese Datei)", "17.06.2026", "Bauherr",
     f"{DROP_ROOT}/Bad-Renovation Locarno - Master.xlsx"),
    ("Lista_artigiani-Handwerkerliste 2024.pdf", "Condominio-Info", "12.08.2024", "Verwaltung Le Palme",
     f"{DROP_ROOT}/Lista_artigiani-Handwerkerliste 2024.pdf"),
    ("Ristrutturazione appartamento no. 10_MIRAGLIA-Regole generali.pdf", "Hausordnung Renovation", "16.10.2023", "Verwaltung Le Palme",
     f"{DROP_ROOT}/Ristrutturazione appartamento no. 10_MIRAGLIA-Regole generali_16.10.2023.pdf"),
    ("Condominio LE PALME - Nuovo Ordinamento della casa.pdf", "Hausordnung", "29.04.2023", "Verwaltung Le Palme",
     f"{DROP_ROOT}/Condominio LE PALME - Nuovo Ordinamento della casa 29.04.2023.pdf"),
]
start = r + 1
for name, typ, dt, firma, path in docs:
    r += 1
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=typ)
    ws.cell(row=r, column=3, value=dt)
    ws.cell(row=r, column=4, value=firma)
    pcell = ws.cell(row=r, column=5, value=path)
    # Hyperlink as file://
    from urllib.parse import quote
    href = "file://" + quote(path)
    pcell.hyperlink = href
    pcell.font = LINK_FONT
style_data_rows(ws, start, r, len(hdr))
ws.freeze_panes = "A5"
ws.sheet_view.showGridLines = False

# Reset active sheet to Cover
wb.active = 0

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)
print(f"✅ Master-Excel erstellt: {OUTPUT}")
print(f"   Größe: {OUTPUT.stat().st_size / 1024:.1f} KB")
print(f"   Tabs : {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"   · {s}")
