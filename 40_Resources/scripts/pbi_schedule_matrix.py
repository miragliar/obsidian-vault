#!/usr/bin/env python3
"""
pbi_schedule_matrix.py — READ-ONLY. Build an Excel refresh-schedule heatmap:
one row per dataset/dataflow, hour-of-day columns showing EXPECTED REFRESHES/DAY,
so you can visually spot the congestion and plan moves.

- Scheduled datasets  -> exact schedule times (refreshSchedule).
- API/Automate/Dataflows (no schedule) -> AVERAGE slots derived from refresh history
  (datasets: /refreshes, dataflows: /transactions) over the last N days.

Reuses pbi_inventory.json for gateway/PROFFIX/schedule, fetches only history here.
Times shown in CEST (UTC+2). Output: .xlsx with 'Heatmap' + 'Overview' sheets.

Usage:
  ./.venv/bin/python pbi_schedule_matrix.py --tenant directhandlingch.onmicrosoft.com \
      --inv /tmp/pbi_inventory.json --out "/path/SHARKGROUP_Refresh_Matrix.xlsx" --days 14
"""
import argparse, json, os, re, sys, time
from collections import defaultdict
from datetime import datetime, timedelta, timezone
import msal, requests
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CID = "04b07795-8ddb-461a-bbee-02f9e1bf7b46"
SCOPE = ["https://analysis.windows.net/powerbi/api/.default"]
BASE = "https://api.powerbi.com/v1.0/myorg"
TZ = timezone(timedelta(hours=2))  # CEST


def log(*a): print(*a, file=sys.stderr, flush=True)


def token(tenant):
    # Token-Cache im macOS Keychain (verschlüsselt, kein Dropbox-Sync) — via auth_common.
    from auth_common import build_pbi_cache
    app = msal.PublicClientApplication(CID, authority=f"https://login.microsoftonline.com/{tenant}", token_cache=build_pbi_cache(tenant))
    acc = app.get_accounts()
    r = app.acquire_token_silent(SCOPE, account=acc[0]) if acc else None
    if not r:
        fl = app.initiate_device_flow(scopes=SCOPE); log(fl["message"]); r = app.acquire_token_by_device_flow(fl)
    return r["access_token"]


def ptime(s):
    if not s: return None
    s = s.strip().replace("Z", "+00:00"); s = re.sub(r"(\.\d{6})\d*(\+)", r"\1\2", s)
    try: d = datetime.fromisoformat(s)
    except ValueError: return None
    return d.astimezone(timezone.utc) if d.tzinfo else d.replace(tzinfo=timezone.utc)


ap = argparse.ArgumentParser()
ap.add_argument("--tenant", required=True)
ap.add_argument("--inv", default="/tmp/pbi_inventory.json")
ap.add_argument("--out", required=True)
ap.add_argument("--days", type=int, default=14)
a = ap.parse_args()

inv = json.load(open(a.inv))
by_id = {r["id"]: r for r in inv["records"]}
now = datetime.now(timezone.utc)
win = now - timedelta(days=a.days)

s = requests.Session(); s.headers["Authorization"] = f"Bearer {token(a.tenant)}"
def val(path):
    for _ in range(4):
        r = s.get(BASE + path, timeout=60)
        if r.status_code == 429: time.sleep(int(r.headers.get("Retry-After", 20))); continue
        if r.status_code >= 400: return []
        return (r.json() or {}).get("value", [])
    return []

rows = []
groups = val("/groups")
log(f"{len(groups)} workspaces; fetching history (last {a.days}d)…")
for gi, g in enumerate(groups, 1):
    gid, gn = g["id"], g.get("name", "?")
    log(f"[{gi}/{len(groups)}] {gn}")
    for d in val(f"/groups/{gid}/datasets"):
        rec = by_id.get(d.get("id"), {})
        starts = [ptime(x.get("startTime")) for x in val(f"/groups/{gid}/datasets/{d['id']}/refreshes?$top=60")]
        rows.append(("dataset", d.get("id"), d.get("name", "?"), gn, rec, [t for t in starts if t]))
    for f in val(f"/groups/{gid}/dataflows"):
        rec = by_id.get(f.get("objectId"), {})
        starts = [ptime(x.get("startTime")) for x in val(f"/groups/{gid}/dataflows/{f['objectId']}/transactions")]
        rows.append(("dataflow", f.get("objectId"), f.get("name", "?"), gn, rec, [t for t in starts if t]))

# ---- build per-row hour vector (expected runs/day) ----
def hour_vector(typ, rec, starts):
    hv = [0.0] * 24
    sched = rec.get("sched_enabled") and rec.get("times")
    if typ == "dataset" and sched:
        for t in rec["times"]:
            try: hv[int(t[:2])] += 1.0
            except Exception: pass
        return hv, "Schedule", "/".join(rec["times"][:8])
    # else: history-derived average
    recent = [t for t in starts if t >= win]
    if not recent:
        return hv, "API", "—"
    days = max(len({t.astimezone(TZ).date() for t in recent}), 1)
    cnt = [0] * 24
    for t in recent: cnt[t.astimezone(TZ).hour] += 1
    for h in range(24): hv[h] = cnt[h] / days
    top = sorted([h for h in range(24) if hv[h] >= 0.25], key=lambda h: -hv[h])[:5]
    txt = "≈" + ",".join(f"{h:02d}h" for h in sorted(top)) if top else "sporadic"
    return hv, "API/hist", txt

data = []
for typ, iid, name, ws, rec, starts in rows:
    hv, trig, times_txt = hour_vector(typ, rec, starts)
    data.append({
        "id": iid, "type": typ, "name": name, "ws": ws,
        "gw": "Y" if rec.get("gw_bound") else "n",
        "needs_gw": "Y" if rec.get("needs_onprem") else "n",
        "proffix": ",".join(d.replace("PX", "") for d in rec.get("proffix_dbs", [])),
        "owner": rec.get("owner", ""), "trig": trig, "times": times_txt,
        "hv": hv, "perday": round(sum(hv), 2),
    })

# PROFFIX-touching first, then workspace, then name
data.sort(key=lambda r: (r["proffix"] == "", r["ws"].lower(), r["name"].lower()))

# ============================ Excel ============================
wb = Workbook()
BOLD = Font(bold=True); WHITEB = Font(bold=True, color="FFFFFF")
HEAD = PatternFill("solid", fgColor="305496"); CENTER = Alignment("center", "center")
PROFFILL = PatternFill("solid", fgColor="FCE4D6")
thin = Side(style="thin", color="D9D9D9"); BORDER = Border(thin, thin, thin, thin)

# ---- Heatmap sheet ----
hm = wb.active; hm.title = "Heatmap"
cols = ["Type", "Name", "Workspace", "GW", "PROFFIX", "Trigger", "/day"] + [f"{h:02d}" for h in range(24)]
hm.append(cols)
for c in range(1, len(cols) + 1):
    cell = hm.cell(1, c); cell.font = WHITEB; cell.fill = HEAD; cell.alignment = CENTER; cell.border = BORDER
H0 = 8  # first hour column index (H)
for r in data:
    row = [r["type"][:2], r["name"], r["ws"], r["gw"], r["proffix"], r["trig"], r["perday"]] + \
          [round(v, 2) if v else None for v in r["hv"]]
    hm.append(row)
    rr = hm.max_row
    if r["proffix"]:
        hm.cell(rr, 5).fill = PROFFILL
last = hm.max_row
# totals
def totrow(label, predicate):
    tot = [0.0] * 24
    for r in data:
        if predicate(r):
            for h in range(24): tot[h] += r["hv"][h]
    hm.append([label, "", "", "", "", "", round(sum(tot), 1)] + [round(v, 2) if v else None for v in tot])
    for c in range(1, 8): hm.cell(hm.max_row, c).font = BOLD
hm.append([])
totrow("TOTAL — all items", lambda r: True)
totrow("TOTAL — PROFFIX only", lambda r: bool(r["proffix"]))

# color scales
rng = f"{get_column_letter(H0)}2:{get_column_letter(H0+23)}{last}"
hm.conditional_formatting.add(rng, ColorScaleRule(
    start_type="num", start_value=0, start_color="FFFFFF",
    mid_type="percentile", mid_value=80, mid_color="FFD666",
    end_type="max", end_color="F8696B"))
trng = f"{get_column_letter(H0)}{last+2}:{get_column_letter(H0+23)}{last+3}"
hm.conditional_formatting.add(trng, ColorScaleRule(
    start_type="num", start_value=0, start_color="FFFFFF",
    mid_type="percentile", mid_value=60, mid_color="9DC3E6",
    end_type="max", end_color="2E75B6"))

hm.freeze_panes = "H2"
widths = [6, 34, 16, 5, 14, 9, 6] + [4] * 24
for i, w in enumerate(widths, 1): hm.column_dimensions[get_column_letter(i)].width = w
for h in range(24):
    hm.cell(1, H0 + h).alignment = CENTER
hm.auto_filter.ref = f"A1:G{last}"

# ---- Overview sheet ----
ov = wb.create_sheet("Overview")
ocols = ["ID", "Type", "Name", "Workspace", "Owner", "Gateway", "NeedsGW", "PROFFIX DBs", "Trigger", "Schedule/Typical times", "Refreshes/day"]
ov.append(ocols)
for c in range(1, len(ocols) + 1):
    cell = ov.cell(1, c); cell.font = WHITEB; cell.fill = HEAD; cell.alignment = CENTER
for r in data:
    ov.append([r["id"], r["type"], r["name"], r["ws"], r["owner"], r["gw"], r["needs_gw"],
               r["proffix"], r["trig"], r["times"], r["perday"]])
    if r["gw"] == "Y" and r["needs_gw"] == "n":
        ov.cell(ov.max_row, 6).fill = PatternFill("solid", fgColor="FFF2CC")  # cloud-only→GW flag
ow = [38, 9, 34, 16, 26, 8, 8, 18, 10, 28, 12]
for i, w in enumerate(ow, 1): ov.column_dimensions[get_column_letter(i)].width = w
ov.freeze_panes = "A2"; ov.auto_filter.ref = ov.dimensions

# ---- Legend sheet ----
lg = wb.create_sheet("Legend")
for i, line in enumerate([
    "SHARKGROUP — Power BI Refresh Schedule Matrix",
    f"Generated {now.astimezone(TZ):%Y-%m-%d %H:%M} CEST · history window: last {a.days} days · times in CEST (UTC+2)",
    "",
    "Heatmap cells = EXPECTED REFRESHES PER DAY in that hour-of-day.",
    "  • Scheduled datasets: from the refresh schedule (exact daily slots).",
    "  • API/Automate & dataflows: AVERAGE per-day runs derived from refresh history.",
    "Red = busy hour. Bottom blue rows = TOTAL load/hour (all items, and PROFFIX-only).",
    "PROFFIX column = on-prem ERP databases the item reads (these cause SQL load on SHAR-SRV-08).",
    "GW=Y bound to on-prem gateway; NeedsGW=n + GW=Y (yellow) = cloud-only, gateway binding removable.",
    "Goal: flatten the PROFFIX-only total curve out of business hours (08:00–18:00).",
], 1):
    lg.cell(i, 1, line)
    if i == 1: lg.cell(i, 1).font = Font(bold=True, size=13)
lg.column_dimensions["A"].width = 110

wb.save(a.out)
log(f"Wrote {a.out}  ({len(data)} items)")
print(f"OK: {len(data)} items -> {a.out}")
